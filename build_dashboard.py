import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

print("--- Phase 4: Building Excel Executive Dashboard (10 Charts) ---")

# File paths
DATA_DIR = r"c:\Users\Lenovo\Desktop\PROJECT 01 E-Commerce Sales Analysis"
files = {
    'orders': os.path.join(DATA_DIR, 'olist_orders_dataset.csv'),
    'items': os.path.join(DATA_DIR, 'olist_order_items_dataset.csv'),
    'products': os.path.join(DATA_DIR, 'olist_products_dataset.csv'),
    'reviews': os.path.join(DATA_DIR, 'olist_order_reviews_dataset.csv'),
    'customers': os.path.join(DATA_DIR, 'olist_customers_dataset.csv'),
    'payments': os.path.join(DATA_DIR, 'olist_order_payments_dataset.csv'),
    'translation': os.path.join(DATA_DIR, 'product_category_name_translation.csv')
}

# 1. Load and clean data (same logic as analysis script)
df_orders = pd.read_csv(files['orders'])
df_items = pd.read_csv(files['items'])
df_products = pd.read_csv(files['products'])
df_reviews = pd.read_csv(files['reviews'])
df_customers = pd.read_csv(files['customers'])
df_payments = pd.read_csv(files['payments'])
df_translation = pd.read_csv(files['translation'], encoding='utf-8-sig')

# Filter orders
active_orders = df_orders[~df_orders['order_status'].isin(['canceled', 'unavailable'])].copy()
for col in ['order_purchase_timestamp', 'order_approved_at', 'order_delivered_carrier_date', 'order_delivered_customer_date', 'order_estimated_delivery_date']:
    active_orders[col] = pd.to_datetime(active_orders[col])

start_date = pd.to_datetime('2017-09-01 00:00:00')
end_date = pd.to_datetime('2018-08-31 23:59:59')
df_orders_12m = active_orders[
    (active_orders['order_purchase_timestamp'] >= start_date) & 
    (active_orders['order_purchase_timestamp'] <= end_date)
].copy()

df_reviews_unique = df_reviews.groupby('order_id', as_index=False).agg({'review_score': 'mean'})
df_translation.columns = df_translation.columns.str.replace(r'^\ufeff', '', regex=True)

# Merge
merged = df_orders_12m.merge(df_items, on='order_id', how='inner')
merged = merged.merge(df_products, on='product_id', how='left')
merged = merged.merge(df_translation, on='product_category_name', how='left')
merged['product_category_name_english'] = merged['product_category_name_english'].fillna('Unknown')
missing_trans = (merged['product_category_name_english'] == 'Unknown') & (merged['product_category_name'].notna())
merged.loc[missing_trans, 'product_category_name_english'] = merged.loc[missing_trans, 'product_category_name'].apply(
    lambda x: str(x).replace('_', ' ').title()
)
merged = merged.merge(df_customers, on='customer_id', how='left')
merged = merged.merge(df_reviews_unique, on='order_id', how='left')
merged['review_score'] = merged['review_score'].fillna(merged['review_score'].mean())
merged['purchase_month'] = merged['order_purchase_timestamp'].dt.strftime('%Y-%m')

# Aggregate KPIs
total_revenue = merged['price'].sum()
total_orders = merged['order_id'].nunique()
aov_overall = total_revenue / total_orders

category_revenue = merged.groupby('product_category_name_english')['price'].sum().sort_values(ascending=False)
top_category = category_revenue.index[0]
top_category_rev = category_revenue.iloc[0]

monthly_sales = merged.groupby('purchase_month')['price'].sum().sort_values(ascending=False)
peak_month = monthly_sales.index[0]
peak_month_sales = monthly_sales.iloc[0]

# Additional aggregations for tables
region_sales = merged.groupby('customer_state')['price'].sum().sort_values(ascending=False)
df_order_reviews = merged.drop_duplicates('order_id')
review_dist = df_order_reviews['review_score'].round().value_counts().sort_index()

# Payment type aggregation
orders_pay = df_orders_12m.merge(df_payments, on='order_id', how='inner')
pay_dist = orders_pay.groupby('payment_type')['payment_value'].sum().sort_values(ascending=False)
pay_dist = pay_dist[pay_dist.index != 'not_defined']
total_pay_val = pay_dist.sum()

# Delivery time aggregation vs reviews
df_delivery = df_orders_12m.dropna(subset=['order_delivered_customer_date']).copy()
df_delivery['delivery_lead_time_days'] = (df_delivery['order_delivered_customer_date'] - df_delivery['order_purchase_timestamp']).dt.days
df_delivery = df_delivery[df_delivery['delivery_lead_time_days'] >= 0]
df_delivery_reviews = df_delivery.merge(df_reviews_unique, on='order_id', how='inner')
delivery_time_vs_review = df_delivery_reviews.groupby('review_score')['delivery_lead_time_days'].mean()

# Create Excel Workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Executive Dashboard"
ws.views.sheetView[0].showGridLines = True

# Font and Styling constants
FONT_NAME = 'Calibri'
COLOR_DARK_BLUE = '1F4E78'
COLOR_LIGHT_BLUE = 'D9E1F2'
COLOR_GRAY = 'F2F2F2'
COLOR_TEXT_LIGHT = 'FFFFFF'

# Styles
font_title = Font(name=FONT_NAME, size=18, bold=True, color=COLOR_TEXT_LIGHT)
font_subtitle = Font(name=FONT_NAME, size=11, italic=True, color=COLOR_TEXT_LIGHT)
font_section = Font(name=FONT_NAME, size=13, bold=True, color=COLOR_DARK_BLUE)
font_header = Font(name=FONT_NAME, size=11, bold=True, color=COLOR_TEXT_LIGHT)
font_kpi_val = Font(name=FONT_NAME, size=16, bold=True, color=COLOR_DARK_BLUE)
font_kpi_lbl = Font(name=FONT_NAME, size=9, bold=True, color='595959')
font_bold = Font(name=FONT_NAME, size=11, bold=True)
font_regular = Font(name=FONT_NAME, size=11)

fill_banner = PatternFill(start_color=COLOR_DARK_BLUE, end_color=COLOR_DARK_BLUE, fill_type='solid')
fill_header = PatternFill(start_color=COLOR_DARK_BLUE, end_color=COLOR_DARK_BLUE, fill_type='solid')
fill_kpi_bg = PatternFill(start_color=COLOR_LIGHT_BLUE, end_color=COLOR_LIGHT_BLUE, fill_type='solid')
fill_zebra = PatternFill(start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type='solid')

align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_left = Alignment(horizontal='left', vertical='center')
align_right = Alignment(horizontal='right', vertical='center')

border_thin = Side(border_style="thin", color="D9D9D9")
border_double = Side(border_style="double", color="000000")
border_thick = Side(border_style="medium", color="1F4E78")

box_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
total_border = Border(top=border_thin, bottom=border_double)

# -------------------------------------------------------------
# 1. Header Banner (A1:P3)
# -------------------------------------------------------------
ws.merge_cells('A1:P1')
ws.merge_cells('A2:P2')
ws.merge_cells('A3:P3')

for row in ws['A1:P3']:
    for cell in row:
        cell.fill = fill_banner

ws['A1'] = "OLIST E-COMMERCE EXECUTIVE PERFORMANCE DASHBOARD"
ws['A1'].font = font_title
ws['A1'].alignment = align_center

ws['A3'] = "Analysis Period: September 2017 – August 2018 (12 Months of Sales Data)"
ws['A3'].font = font_subtitle
ws['A3'].alignment = align_center

ws.row_dimensions[1].height = 25
ws.row_dimensions[2].height = 10
ws.row_dimensions[3].height = 20

# -------------------------------------------------------------
# 2. KPI Cards (Row 5 to 7)
# -------------------------------------------------------------
def create_kpi_card(ws, start_col, label, value, subtext):
    col1 = start_col
    col2 = chr(ord(start_col) + 1)
    
    ws.merge_cells(f"{col1}5:{col2}5")
    ws.merge_cells(f"{col1}6:{col2}6")
    ws.merge_cells(f"{col1}7:{col2}7")
    
    c_lbl = ws[f"{col1}5"]
    c_val = ws[f"{col1}6"]
    c_sub = ws[f"{col1}7"]
    
    c_lbl.value = label
    c_lbl.font = font_kpi_lbl
    c_lbl.alignment = align_center
    
    c_val.value = value
    c_val.font = font_kpi_val
    c_val.alignment = align_center
    
    c_sub.value = subtext
    c_sub.font = font_kpi_lbl
    c_sub.alignment = align_center
    
    for r in range(5, 8):
        for c in [col1, col2]:
            cell = ws[f"{c}{r}"]
            cell.fill = fill_kpi_bg
            left_side = border_thick if c == col1 else None
            right_side = border_thick if c == col2 else None
            top_side = border_thick if r == 5 else None
            bottom_side = border_thick if r == 7 else None
            cell.border = Border(left=left_side, right=right_side, top=top_side, bottom=bottom_side)

create_kpi_card(ws, 'A', "TOTAL REVENUE", f"R$ {total_revenue:,.2f}", "Gross product sales")
create_kpi_card(ws, 'D', "TOTAL ORDERS", f"{total_orders:,}", "Completed orders count")
create_kpi_card(ws, 'G', "AVG ORDER VALUE (AOV)", f"R$ {aov_overall:.2f}", "Revenue per order")
create_kpi_card(ws, 'J', "TOP PRODUCT CATEGORY", top_category.replace('_', ' ').title(), f"R$ {top_category_rev:,.2f} sales")
create_kpi_card(ws, 'M', "PEAK SALES MONTH", f"{peak_month}", f"R$ {peak_month_sales:,.2f} revenue")

ws.row_dimensions[5].height = 15
ws.row_dimensions[6].height = 25
ws.row_dimensions[7].height = 15

# -------------------------------------------------------------
# 3. Section Banner (Row 9)
# -------------------------------------------------------------
ws['A9'] = "PERFORMANCE ANALYSIS TABLES"
ws['A9'].font = font_section
ws.row_dimensions[9].height = 22

# -------------------------------------------------------------
# 4. Data Tables
# -------------------------------------------------------------

# Table A: Monthly Trend (Cols A-D)
ws['A11'] = "Month"
ws['B11'] = "Revenue"
ws['C11'] = "Orders"
ws['D11'] = "AOV"

monthly_sorted = merged.groupby('purchase_month').agg(
    revenue=('price', 'sum'),
    orders=('order_id', 'nunique')
).sort_index()
monthly_sorted['aov'] = monthly_sorted['revenue'] / monthly_sorted['orders']

row_idx = 12
for m, r in monthly_sorted.iterrows():
    ws.cell(row=row_idx, column=1, value=m).alignment = align_center
    ws.cell(row=row_idx, column=2, value=r['revenue'])
    ws.cell(row=row_idx, column=3, value=r['orders'])
    ws.cell(row=row_idx, column=4, value=r['aov'])
    row_idx += 1

# Monthly Totals row
ws.cell(row=row_idx, column=1, value="Total/Average").font = font_bold
ws.cell(row=row_idx, column=1).alignment = align_center
ws.cell(row=row_idx, column=2, value=f"=SUM(B12:B{row_idx-1})").font = font_bold
ws.cell(row=row_idx, column=3, value=f"=SUM(C12:C{row_idx-1})").font = font_bold
ws.cell(row=row_idx, column=4, value=f"=B{row_idx}/C{row_idx}").font = font_bold
table_a_last_row = row_idx

# Format Table A
for r in range(11, table_a_last_row + 1):
    for c in range(1, 5):
        cell = ws.cell(row=r, column=c)
        if r == 11:
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
        elif r == table_a_last_row:
            cell.border = total_border
        else:
            cell.font = font_regular
            cell.border = box_border
            if r % 2 == 1:
                cell.fill = fill_zebra
        if r > 11:
            if c in [2, 4]:
                cell.number_format = '"R$"#,##0.00'
                cell.alignment = align_right
            elif c == 3:
                cell.number_format = '#,##0'
                cell.alignment = align_right


# Table B: Top 10 Categories by Revenue (Cols F-H)
ws['F11'] = "Top Category"
ws['G11'] = "Revenue"
ws['H11'] = "% Share"

row_idx = 12
for cat, rev in category_revenue.head(10).items():
    ws.cell(row=row_idx, column=6, value=cat.replace('_', ' ').title()).alignment = align_left
    ws.cell(row=row_idx, column=7, value=rev)
    ws.cell(row=row_idx, column=8, value=f"=G{row_idx}/{total_revenue}").alignment = align_right
    row_idx += 1

# Table B Totals
ws.cell(row=row_idx, column=6, value="Top 10 Total").font = font_bold
ws.cell(row=row_idx, column=7, value=f"=SUM(G12:G{row_idx-1})").font = font_bold
ws.cell(row=row_idx, column=8, value=f"=SUM(H12:H{row_idx-1})").font = font_bold
table_b_last_row = row_idx

# Format Table B
for r in range(11, table_b_last_row + 1):
    for c in range(6, 9):
        cell = ws.cell(row=r, column=c)
        if r == 11:
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
        elif r == table_b_last_row:
            cell.border = total_border
        else:
            cell.font = font_regular
            cell.border = box_border
            if r % 2 == 1:
                cell.fill = fill_zebra
        if r > 11:
            if c == 7:
                cell.number_format = '"R$"#,##0.00'
                cell.alignment = align_right
            elif c == 8:
                cell.number_format = '0.0%'
                cell.alignment = align_right


# Table C: Regional Sales (Cols J-L)
ws['J11'] = "State"
ws['K11'] = "Revenue"
ws['L11'] = "% Share"

row_idx = 12
for st, rev in region_sales.head(10).items():
    ws.cell(row=row_idx, column=10, value=st).alignment = align_center
    ws.cell(row=row_idx, column=11, value=rev)
    ws.cell(row=row_idx, column=12, value=f"=K{row_idx}/{total_revenue}").alignment = align_right
    row_idx += 1

# Table C Totals
ws.cell(row=row_idx, column=10, value="Top 10 Total").font = font_bold
ws.cell(row=row_idx, column=10).alignment = align_center
ws.cell(row=row_idx, column=11, value=f"=SUM(K12:K{row_idx-1})").font = font_bold
ws.cell(row=row_idx, column=12, value=f"=SUM(L12:L{row_idx-1})").font = font_bold
table_c_last_row = row_idx

# Format Table C
for r in range(11, table_c_last_row + 1):
    for c in range(10, 13):
        cell = ws.cell(row=r, column=c)
        if r == 11:
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
        elif r == table_c_last_row:
            cell.border = total_border
        else:
            cell.font = font_regular
            cell.border = box_border
            if r % 2 == 1:
                cell.fill = fill_zebra
        if r > 11:
            if c == 11:
                cell.number_format = '"R$"#,##0.00'
                cell.alignment = align_right
            elif c == 12:
                cell.number_format = '0.0%'
                cell.alignment = align_right


# Table D: Customer Review Score Distribution (Cols N-P)
ws['N11'] = "Review Score"
ws['O11'] = "Orders"
ws['P11'] = "% Share"

row_idx = 12
for score, val in review_dist.items():
    ws.cell(row=row_idx, column=14, value=f"{int(score)} Stars").alignment = align_center
    ws.cell(row=row_idx, column=15, value=val)
    ws.cell(row=row_idx, column=16, value=f"=O{row_idx}/{df_order_reviews.shape[0]}").alignment = align_right
    row_idx += 1

# Table D Totals
ws.cell(row=row_idx, column=14, value="Total").font = font_bold
ws.cell(row=row_idx, column=14).alignment = align_center
ws.cell(row=row_idx, column=15, value=f"=SUM(O12:O{row_idx-1})").font = font_bold
ws.cell(row=row_idx, column=16, value=f"=SUM(P12:P{row_idx-1})").font = font_bold
table_d_last_row = row_idx

# Format Table D
for r in range(11, table_d_last_row + 1):
    for c in range(14, 17):
        cell = ws.cell(row=r, column=c)
        if r == 11:
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
        elif r == table_d_last_row:
            cell.border = total_border
        else:
            cell.font = font_regular
            cell.border = box_border
            if r % 2 == 1:
                cell.fill = fill_zebra
        if r > 11:
            if c == 15:
                cell.number_format = '#,##0'
                cell.alignment = align_right
            elif c == 16:
                cell.number_format = '0.0%'
                cell.alignment = align_right


# --- NEW TABLES FOR NEW METRICS (Rows 25 to 31) ---

# Table E: Payment Types Distribution (Cols F-H, Rows 25 to 31)
ws['F25'] = "Payment Type"
ws['G25'] = "Payment Value"
ws['H25'] = "% Share"

row_idx = 26
for pt, val in pay_dist.items():
    ws.cell(row=row_idx, column=6, value=pt.replace('_', ' ').title()).alignment = align_left
    ws.cell(row=row_idx, column=7, value=val)
    ws.cell(row=row_idx, column=8, value=f"=G{row_idx}/{total_pay_val}").alignment = align_right
    row_idx += 1

# Table E Totals
ws.cell(row=row_idx, column=6, value="Total Payments").font = font_bold
ws.cell(row=row_idx, column=7, value=f"=SUM(G26:G{row_idx-1})").font = font_bold
ws.cell(row=row_idx, column=8, value=f"=SUM(H26:H{row_idx-1})").font = font_bold
table_e_last_row = row_idx

# Format Table E
for r in range(25, table_e_last_row + 1):
    for c in range(6, 9):
        cell = ws.cell(row=r, column=c)
        if r == 25:
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
        elif r == table_e_last_row:
            cell.border = total_border
        else:
            cell.font = font_regular
            cell.border = box_border
            if r % 2 == 1:
                cell.fill = fill_zebra
        if r > 25:
            if c == 7:
                cell.number_format = '"R$"#,##0.00'
                cell.alignment = align_right
            elif c == 8:
                cell.number_format = '0.0%'
                cell.alignment = align_right


# Table F: Delivery Time vs. Satisfaction (Cols J-L, Rows 25 to 31)
ws['J25'] = "Review Rating"
ws['K25'] = "Avg Delivery Time"
ws['L25'] = "Target (Days)"

row_idx = 26
for rating, days in delivery_time_vs_review.items():
    ws.cell(row=row_idx, column=10, value=f"{int(rating)} Stars").alignment = align_center
    ws.cell(row=row_idx, column=11, value=days)
    ws.cell(row=row_idx, column=12, value=12 if rating >= 4 else 8).alignment = align_right # target SLA
    row_idx += 1

# Table F Averages
ws.cell(row=row_idx, column=10, value="Overall Average").font = font_bold
ws.cell(row=row_idx, column=10).alignment = align_center
ws.cell(row=row_idx, column=11, value=f"=AVERAGE(K26:K{row_idx-1})").font = font_bold
ws.cell(row=row_idx, column=12, value="").font = font_bold
table_f_last_row = row_idx

# Format Table F
for r in range(25, table_f_last_row + 1):
    for c in range(10, 13):
        cell = ws.cell(row=r, column=c)
        if r == 25:
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
        elif r == table_f_last_row:
            cell.border = total_border
        else:
            cell.font = font_regular
            cell.border = box_border
            if r % 2 == 1:
                cell.fill = fill_zebra
        if r > 25:
            if c == 11:
                cell.number_format = '0.0" Days"'
                cell.alignment = align_right
            elif c == 12 and cell.value != "":
                cell.number_format = '0" Days"'
                cell.alignment = align_right

# Align heights of rows 25 to 32
for r in range(25, 33):
    ws.row_dimensions[r].height = 18

# -------------------------------------------------------------
# 5. Insert Visualizations (Row 34 onwards)
# -------------------------------------------------------------
ws['A34'] = "PERFORMANCE VISUALISATIONS & CHARTS (10 CHARTS REPORT)"
ws['A34'].font = font_section
ws.row_dimensions[34].height = 22

# Embed custom generated high-quality charts in a 2-column grid layout
chart_positions = [
    ("visualizations/02_monthly_sales_trend.png", "A36", "Monthly Revenue & Order Volume Trends"),
    ("visualizations/01_revenue_by_category.png", "I36", "Top 10 Product Categories by Revenue"),
    ("visualizations/03_regional_sales.png", "A56", "Regional Sales Distribution by State"),
    ("visualizations/05_review_scores_distribution.png", "I56", "Customer Satisfaction (Review Scores)"),
    ("visualizations/04_order_values_distribution.png", "A76", "Order Value Histogram & Distribution"),
    ("visualizations/06_category_monthly_heatmap.png", "I76", "Monthly Revenue Heatmap (Top Categories)"),
    ("visualizations/07_payment_types_distribution.png", "A96", "Payment Method Value Distributions"),
    ("visualizations/08_order_volume_by_day_of_week.png", "I96", "Order Volumes by Day of the Week"),
    ("visualizations/09_average_freight_by_state.png", "A116", "Highest Average Shipping Cost by State"),
    ("visualizations/10_delivery_time_vs_review_score.png", "I116", "Delivery Time in Days vs Review Scores")
]

for path, cell, title in chart_positions:
    try:
        img = Image(path)
        img.width = 600
        img.height = 360
        ws.add_image(img, cell)
        print(f"Embedded Chart '{title}' at {cell}.")
    except Exception as e:
        print(f"Error embedding chart '{title}': {e}")

# Row heights for chart areas
for r in range(36, 140, 20):
    ws.row_dimensions[r].height = 20

# -------------------------------------------------------------
# 6. Auto-fit column widths
# -------------------------------------------------------------
column_widths = {
    'A': 12, 'B': 18, 'C': 12, 'D': 15,  # Table A
    'E': 4,                              # Spacer
    'F': 25, 'G': 18, 'H': 12,           # Table B / Table E
    'I': 4,                              # Spacer
    'J': 16, 'K': 18, 'L': 15,           # Table C / Table F
    'M': 4,                              # Spacer
    'N': 16, 'O': 12, 'P': 12            # Table D
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Save the workbook
output_path = os.path.join(DATA_DIR, "dashboard.xlsx")
wb.save(output_path)
print(f"Successfully generated executive dashboard with 10 charts at: {output_path}")
